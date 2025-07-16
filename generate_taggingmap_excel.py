import sys
import json
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import requests
from io import BytesIO

def generate_taggingmap_excel(
    excel_filename,
    info_text_block,     # 리스트: 상단에 넣을 여러 줄 텍스트
    table_columns,       # 리스트: 동적으로 변하는 표 헤더
    table_rows,          # 리스트[dict]: 각 표 row {col: value, ...}
    image_urls,          # 이미지 url 리스트, 표 데이터와 순서 매칭
    image_col_width=80,  # A열(이미지) 너비 px
    image_max_width=80,  # 이미지 최대 넓이(px)
    image_max_height=80  # 이미지 최대 높이(px)
):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "태깅맵"

    row_offset = 1
    # 1. 상단 info block 삽입 (각 한 셀에 한 줄)
    for line in info_text_block:
        ws.cell(row=row_offset, column=1, value=line)
        row_offset += 1
    row_offset += 1  # 한 줄 띄움

    # 2. 표 헤더 작성 (A열은 "이미지")
    header_row = row_offset
    ws.cell(row=header_row, column=1, value="이미지")
    for idx, col in enumerate(table_columns):
        ws.cell(row=header_row, column=2+idx, value=col)

    # 3. 데이터 행 작성
    for ridx, row in enumerate(table_rows):
        excel_row = header_row + 1 + ridx
        # 이미지 먼저
        if image_urls and ridx < len(image_urls) and image_urls[ridx]:
            try:
                resp = requests.get(image_urls[ridx])
                img = XLImage(BytesIO(resp.content))
                # 비율 맞춰서 크기 제한
                ratio = min(image_max_width / img.width, image_max_height / img.height, 1.0)
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
                img.anchor = f"A{excel_row}"  # 엑셀 셀 지정
                ws.add_image(img)
            except Exception as e:
                ws.cell(row=excel_row, column=1, value="[이미지 오류]")
        # 데이터 (동적 컬럼)
        for cidx, col in enumerate(table_columns):
            value = row.get(col, "")
            ws.cell(row=excel_row, column=2 + cidx, value=value)
            ws.cell(row=excel_row, column=2 + cidx).alignment = Alignment(vertical="center")

    # 4. A열 너비 고정
    ws.column_dimensions['A'].width = image_col_width / 7  # openpyxl은 "문자 폭" 기준, 대략 7px=1

    # 5. 나머지 열 너비/정렬
    for idx, col in enumerate(table_columns):
        col_letter = get_column_letter(2 + idx)
        ws.column_dimensions[col_letter].width = max(15, len(col)*2)  # 적당히 자동

    # 6. 저장
    wb.save(excel_filename)
    print(f"엑셀 파일 저장 완료: {excel_filename}")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python generate_taggingmap_excel.py input.json output.xlsx")
        sys.exit(1)
    input_json = sys.argv[1]
    output_excel = sys.argv[2]
    with open(input_json, encoding="utf-8") as f:
        data = json.load(f)
    info_block = data.get("info", [])
    columns = data.get("columns", [])
    rows = data.get("rows", [])
    images = data.get("images", [])
    generate_taggingmap_excel(
        output_excel,
        info_block,
        columns,
        rows,
        images
    )